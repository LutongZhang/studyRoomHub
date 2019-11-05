import pandas as pd
from openpyxl import load_workbook


class ExcelOperator:
    def __init__(self):
        # self.excel = openpyxl.load_workbook('Template.xlsx')
        # #(time - 8)/0.5+5
        # self.indexToColumn = {1: "C", 2: "D",
        #                       3: "E", 4: "f", 5: "G", 6: "H", 7: "I"}
        self.fn = "Template.xlsx"

    def fillCell(self, number, startTime, roomName):
        row = int((startTime - 8)/0.5+5)
        sheetName = "Sheet1"
        if number > 7:
            number -= 7
            sheetName = "Sheet2"

        df = pd.read_excel(self.fn, header=None, sheet_name=sheetName)
        writer = pd.ExcelWriter(self.fn, engine='openpyxl')
        book = load_workbook(self.fn)

        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df2 = pd.DataFrame({'Data': [roomName, roomName, roomName, roomName]})
        df.to_excel(writer, sheet_name=sheetName, header=None, index=False)
        df2.to_excel(writer, sheet_name=sheetName, header=None, index=False,
                     startcol=number+1, startrow=row-1)
        writer.save()
